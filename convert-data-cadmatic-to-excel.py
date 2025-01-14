import os
import re
import openpyxl
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from tkinter import Tk, filedialog, Button, Label, StringVar, Frame, messagebox, PhotoImage

# Unified process_file function to handle profile type and bar number detection
def process_file(file_path):
    profile_types = []
    bar_numbers = []

    with open(file_path, 'r') as file:
        current_profile = None
        for line in file:
            line = line.replace('|', '').strip()

            # Detect Profile type
            profile_match = re.search(r"Profile type\s*:\s*(\S+)", line)
            if profile_match:
                current_profile = profile_match.group(1)

            # Detect Bar number and append both Profile type and Bar number
            bar_number_match = re.search(r"Bar number\s*:\s*(\d+)", line)
            if bar_number_match and current_profile:
                profile_types.append(current_profile)
                bar_numbers.append(int(bar_number_match.group(1)))

    # Aggregate by Profile type and count the occurrences of Bar number
    profile_summary = {}
    for profile, bar in zip(profile_types, bar_numbers):
        if profile not in profile_summary:
            profile_summary[profile] = []
        profile_summary[profile].append(bar)

    # Create the resulting DataFrame with aggregated counts
    profile_type_resumes = []
    bar_number_resumes = []
    for profile, bars in profile_summary.items():
        profile_type_resumes.append(profile)
        bar_number_resumes.append(len(bars))  # Count occurrences of each profile type

    df = pd.DataFrame({
        "Profile type resume": profile_type_resumes,
        "Bar number resume": bar_number_resumes
    })
    
    return df

# Function to split the Profile type into separate columns
def split_profile_type(profile_type):
    # match = re.match(r"([A-Za-z]+)(\d+)((X\d+)*)", profile_type)
    
    # if match:
    #     letters = match.group(1)
    #     numbers = [match.group(2)]
        
    #     if match.group(3):
    #         numbers.extend(match.group(3).split('X')[1:])
        
    #     return [letters] + numbers
    # else:
    #     return [profile_type]

     # Regex untuk memisahkan huruf, angka, dan X
    pattern = r'[A-Za-z]+|\d+|X'
    matches = re.findall(pattern, profile_type)
    
    # Pastikan hasilnya memiliki 8 elemen dengan nilai default
    default_values = ['X', '0', 'X', '0', 'X', '0', 'X', '0']
    matches.extend(default_values[len(matches):])
    
    return matches[:8]

# Function to parse the .list file
def parse_list_file(file_path):
    data = {
        "Profile type": [],
        "Bar-codenr": [],
        "Length bar": [],
        "Material": [],
        "Bar number": [],
        "Total length": [],
        "Scrap-iron": [],
        "Part": [],
        "Cut off Length": []
    }

    header_data = {
        "Object": "",
        "Block": "",
        "Date": ""
    }

    with open(file_path, 'r') as file:
        part_section = False
        common_data = {}
        for line in file:
            line = line.replace('|', '').strip()

            object_match = re.search(r"Object:\s*(\d+)", line)
            block_match = re.search(r"Block:\s*(\d+)", line)
            date_match = re.search(r"Date:\s*(\S+)", line)

            if object_match:
                header_data["Object"] = object_match.group(1)
            if block_match:
                header_data["Block"] = block_match.group(1)
            if date_match:
                header_data["Date"] = date_match.group(1)

            if re.match(r"^Part\s+Cut off Length", line):
                part_section = True
                continue

            if part_section and re.match(r"^\d+\s+\d+", line):
                parts = line.split()
                if len(parts) >= 2:
                    data["Part"].append(parts[0])
                    data["Cut off Length"].append(parts[1])
                    for key in common_data:
                        data[key].append(common_data[key])
                continue

            profile_match = re.search(r"Profile type\s*:\s*(.*)", line)
            barcode_match = re.search(r"Bar-codenr\s*:\s*(.*)", line)
            length_bar_match = re.search(r"Length bar\s*:\s*(.*)", line)
            material_match = re.search(r"Material\s*:\s*(.*)", line)
            bar_number_match = re.search(r"Bar number\s*:\s*(.*)", line)
            total_length_match = re.search(r"Total length\s*:\s*(.*)", line)
            scrap_iron_match = re.search(r"Scrap-iron\s*:\s*(.*)", line)

            if profile_match:
                common_data["Profile type"] = profile_match.group(1).split()[0].strip()
            if barcode_match:
                common_data["Bar-codenr"] = barcode_match.group(1).strip()
            if length_bar_match:
                common_data["Length bar"] = length_bar_match.group(1).split()[0].strip()
            if material_match:
                common_data["Material"] = material_match.group(1).split()[0].strip()
            if bar_number_match:
                common_data["Bar number"] = bar_number_match.group(1).strip()
            if total_length_match:
                common_data["Total length"] = total_length_match.group(1).strip()
            if scrap_iron_match:
                common_data["Scrap-iron"] = scrap_iron_match.group(1).strip()

    return data, header_data

# Function to parse the .lst file
def parse_lst_file(lst_file_path):
    lst_data = {}
    with open(lst_file_path, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith('*'):
                parts = line.split('|')
                if len(parts) >= 6:
                    bcode = parts[0].strip('* ')
                    length = parts[5].strip()
                    if length not in lst_data:
                        lst_data[length] = []
                    lst_data[length].append(bcode)
    return lst_data

# Function to load CSV data
def load_csv_data(csv_file_path):
    csv_data = []
    with open(csv_file_path, newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=';', quotechar='"')
        for row in reader:
            csv_data.append(row)
    return csv_data

# Function to convert specific columns to numbers
def convert_to_number(ws, row):
    columns_to_convert = [2, 3, 5, 6, 7, 9, 14, 13, 15, 16, 17, 19]
    for col in columns_to_convert:
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            try:
                cell.value = float(cell.value)
            except ValueError:
                pass

# Function to convert barcode columns
def convert_to_number_barcode(ws, row):
    """Convert specific columns to integers and preserve full number representation."""
    columns_to_convert = [2]  # Kolom yang akan dikonversi
    for col in columns_to_convert:
        cell = ws.cell(row=row, column=col)  # Ambil sel pada baris dan kolom tertentu
        if cell.value is not None:  # Cek apakah sel tidak kosong
            try:
                # Konversi nilai menjadi integer
                value_as_int = int(float(cell.value))
                # Simpan nilai sebagai string untuk menghindari format eksponensial
                cell.value = str(value_as_int)
                # Set format Excel ke teks
                cell.number_format = '@'
            except ValueError:
                pass  # Abaikan jika konversi gagal

# Function to merge data and export to Excel
def convert_list_to_xlsx(list_file_path, lst_file_path, resume_data, csv_file_path, output_file_path):
    data, header_data = parse_list_file(list_file_path)
    lst_data = parse_lst_file(lst_file_path)
    csv_data = load_csv_data(csv_file_path)

    wb = Workbook()
    ws = wb.active

    ws.append(["PROJECT =", header_data["Object"], "BLOCK =", header_data["Block"], "DATE =", header_data["Date"], "", "", "", "", "", "RESUME"])
    # Header baris kedua tanpa "Profile type resume"
    headers = [
        "PROFILE TYPE", "BAR-CODE", "LENGTH BAR", "MAT", "BAR NUMBER",
        "TOT LENGTH", "SCRAP IRON", "PART NAME", "Cut off Length", "", "",
        "Profile Type", "Height", "", "Width", "", "Thick1", "", "Thick2", "Bar number resume"
    ]
    ws.append(headers)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    for cell in ws[2]:
        cell.font = bold_font

    ws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 16
    ws.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 9
    ws.column_dimensions[openpyxl.utils.get_column_letter(8)].width = 21
    ws.column_dimensions[openpyxl.utils.get_column_letter(12)].width = 18
    ws.column_dimensions[openpyxl.utils.get_column_letter(13)].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(20)].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 15
    for col in [3, 6, 7]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    for col in [5, 9]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 13
    for col in range(14, 17):  
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 7

    # Set row heights
    ws.row_dimensions[2].height = 45.75  # Row 2 height
    for row in range(3, ws.max_row + 1):  # Rows 3 to last row
        ws.row_dimensions[row].height = 25

    # Fill data and set alignment
    bcode_tracker = {key: 0 for key in lst_data}
    for i in range(len(data["Part"])):
        part_value = data["Part"][i]
        cut_off_length = data["Cut off Length"][i]

        if cut_off_length in lst_data:
            index = bcode_tracker[cut_off_length] % len(lst_data[cut_off_length])
            part_value = lst_data[cut_off_length][index]
            bcode_tracker[cut_off_length] += 1

        # Proses split_profile_type jika data ada
        if i < len(resume_data["Profile type resume"]) and resume_data["Profile type resume"][i]:
            split_profile = split_profile_type(resume_data["Profile type resume"][i])
            #split_profile.extend([0] * (8 - len(split_profile)))  # Tambahkan nilai 0 jika kurang
        else:
            split_profile = []*8 # Nilai default jika kosong

        row = [
            data["Profile type"][i],
            data["Bar-codenr"][i],
            data["Length bar"][i],
            data["Material"][i],
            data["Bar number"][i],
            data["Total length"][i],
            data["Scrap-iron"][i],
            part_value,
            cut_off_length,
            "",
            "",
        ] + split_profile + [resume_data["Bar number resume"][i] if i < len(resume_data["Bar number resume"]) else ""]
        ws.append(row)

    for row in range(3, ws.max_row + 1):  # Baris 3 hingga terakhir dengan data
        ws.row_dimensions[row].height = 25

    for excel_row in range(3, ws.max_row + 1):
        convert_to_number(ws, excel_row)

    for excel_row in range(3, ws.max_row + 1):
        excel_col_a = ws.cell(row=excel_row, column=1).value
        excel_col_c = ws.cell(row=excel_row, column=3).value
        excel_col_d = ws.cell(row=excel_row, column=4).value

        excel_col_a_str = str(excel_col_a).strip() if excel_col_a is not None else ""
        excel_col_d_str = str(excel_col_d).strip() if excel_col_d is not None else ""
        excel_col_c_int = int(excel_col_c) if isinstance(excel_col_c, float) else excel_col_c
        excel_col_c_str = str(excel_col_c_int).strip() if excel_col_c_int is not None else ""

        for csv_row in range(len(csv_data)):
            if len(csv_data[csv_row]) < 3:
                continue

            csv_col_b = csv_data[csv_row][1].strip()
            csv_col_c = csv_data[csv_row][2].strip()
            csv_col_d = csv_data[csv_row][3].strip()

            if excel_col_a_str == csv_col_b and excel_col_c_str == csv_col_c and excel_col_d_str == csv_col_d:
                ws.cell(row=excel_row, column=2).value = csv_data[csv_row][0]
                convert_to_number_barcode(ws, excel_row)
                break

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if 1 <= cell.column <= 9 or 12 <= cell.column <= 20:  
                cell.alignment = Alignment(horizontal='center', vertical='center')  

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            if cell.column == 1 or cell.column == 12 or cell.column == 8:    
                cell.alignment = Alignment(horizontal='left', vertical='center')  

    # Remove duplicate values in columns A to G by clearing their content
    unique_rows = set()

    for row in range(3, ws.max_row + 1):
        # Collect data from columns A to G
        row_data = tuple(
            ws.cell(row=row, column=col).value for col in range(1, 8)
        )

        # If the row is a duplicate, clear columns A to G
        if row_data in unique_rows:
            for col in range(1, 8):  # Columns A to G
                cell = ws.cell(row=row, column=col)
                cell.value = None  # Clear the cell value
                cell.border = None  # Remove the border
        else:
            unique_rows.add(row_data)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    top_thick_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thick"),
        bottom=Side(style="thin")
    )

    for cell in ws[2]:  # Baris ke-2
        if 1 <= cell.column <= 9: # Kolom A sampai I atau L sampai T
            cell.border = Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )

    # Memberikan border untuk kolom 1–9
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        # Periksa apakah semua kolom A sampai I memiliki data
        if all(ws.cell(row=row[0].row, column=col).value not in [None, "", " "] for col in range(1, 10)):
            # Border tebal untuk seluruh baris jika semua sel dari kolom A sampai I memiliki data
            for cell in row:
                if 1 <= cell.column <= 9:
                    cell.border = top_thick_border
        else:
            # Border tipis untuk sel yang memiliki data
            for cell in row:
                if 1 <= cell.column <= 9 and cell.value not in [None, "", " "]:
                    cell.border = thin_border

    last_row_with_data = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=12, max_col=20):
        if any(cell.value not in [None, "", " "] for cell in row):  # Cek apakah ada data yang valid
            last_row_with_data = row[0].row

    # Memberikan border hanya untuk kolom 12–18 pada baris terakhir dengan data
    if last_row_with_data > 0:  # Pastikan ada data pada kolom 12–18
        for cell in ws.iter_rows(min_row=2, max_row=last_row_with_data, min_col=12, max_col=20):
            for sub_cell in cell:
                sub_cell.border = thin_border

    
    for cell in ws[2]:  # Baris ke-2
        if 12 <= cell.column <= 20:  # Kolom A sampai I atau L sampai T
            cell.border = Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )

    wb.save(output_file_path)

class App:
    def __init__(self, master):
        self.master = master

        # Set the window icon
        try:
            # For Windows
            master.iconbitmap('icon.ico') 
        except Exception as e:
            print(f"Error setting icon: {e}")

        master.title("Partlist Converter")
        master.geometry("700x500")
        master.configure(bg="#e6f7ff")

        self.frame = Frame(master, bg="#e6f7ff")
        self.frame.pack(pady=40)

        self.list_file = StringVar()
        self.lst_file = StringVar()
        self.csv_file = StringVar()
        self.output_file = StringVar()

        self.create_widgets()

    def create_widgets(self):
        label_font = ("Arial", 14, "bold")
        button_font = ("Arial", 12)

        # .list file selection
        Label(self.frame, text="Select .list file        :", bg="#e6f7ff", font=label_font, anchor='w').grid(row=0, column=0, sticky='w', padx=20, pady=10)
        self.list_button = Button(self.frame, text="Browse", command=self.select_list_file, bg="#b30000", fg="white", font=button_font, width=12, height=1)
        self.list_button.grid(row=0, column=1, padx=20, pady=10)
        self.list_file_label = Label(self.frame, text="", bg="#e6f7ff", fg="green", font=("Arial", 10))
        self.list_file_label.grid(row=0, column=2, sticky='w', padx=20)

        # .lst file selection
        Label(self.frame, text="Select .lst file         :", bg="#e6f7ff", font=label_font, anchor='w').grid(row=1, column=0, sticky='w', padx=20, pady=10)
        self.lst_button = Button(self.frame, text="Browse", command=self.select_lst_file, bg="#b30000", fg="white", font=button_font, width=12, height=1)
        self.lst_button.grid(row=1, column=1, padx=20, pady=10)
        self.lst_file_label = Label(self.frame, text="", bg="#e6f7ff", fg="green", font=("Arial", 10))
        self.lst_file_label.grid(row=1, column=2, sticky='w', padx=20)

        # .csv file selection
        Label(self.frame, text="Select .csv file       :", bg="#e6f7ff", font=label_font, anchor='w').grid(row=2, column=0, sticky='w', padx=20, pady=10)
        self.csv_button = Button(self.frame, text="Browse", command=self.select_csv_file, bg="#b30000", fg="white", font=button_font, width=12, height=1)
        self.csv_button.grid(row=2, column=1, padx=20, pady=10)
        self.csv_file_label = Label(self.frame, text="", bg="#e6f7ff", fg="green", font=("Arial", 10))
        self.csv_file_label.grid(row=2, column=2, sticky='w', padx=20)

        # Output file selection
        Label(self.frame, text="Select output file   :", bg="#e6f7ff", font=label_font, anchor='w').grid(row=3, column=0, sticky='w', padx=20, pady=10)
        self.output_button = Button(self.frame, text="Browse", command=self.select_output_file, bg="#b30000", fg="white", font=button_font, width=12, height=1)
        self.output_button.grid(row=3, column=1, padx=20, pady=10)
        self.output_file_label = Label(self.frame, text="", bg="#e6f7ff", fg="green", font=("Arial", 10))
        self.output_file_label.grid(row=3, column=2, sticky='w', padx=20)

        # Convert button
        Button(self.frame, text="Convert", command=self.convert_files, bg="#355887", fg="white", font=("Arial", 14), width=15, height=2).grid(row=4, columnspan=3, pady=30)

    def select_list_file(self):
        file_path = filedialog.askopenfilename(title="Select .list file", filetypes=[("List files", "*.list"), ("All files", "*.*")])
        if file_path:
            self.list_file.set(file_path)
            self.list_file_label.config(text="Loaded")
            self.list_button.config(bg="#358737")

    def select_lst_file(self):
        file_path = filedialog.askopenfilename(title="Select .lst file", filetypes=[("LST files", "*.lst"), ("All files", "*.*")])
        if file_path:
            self.lst_file.set(file_path)
            self.lst_file_label.config(text="Loaded")
            self.lst_button.config(bg="#358737")

    def select_csv_file(self):
        file_path = filedialog.askopenfilename(title="Select .csv file", filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if file_path:
            self.csv_file.set(file_path)
            self.csv_file_label.config(text="Loaded")
            self.csv_button.config(bg="#358737")

    def select_output_file(self):
        file_path = filedialog.asksaveasfilename(title="Save Output File", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path:
            self.output_file.set(file_path)
            self.output_file_label.config(text="Save")
            self.output_button.config(bg="#358737")

    def convert_files(self):
        try:
            resume_df = process_file(self.list_file.get())
            convert_list_to_xlsx(self.list_file.get(), self.lst_file.get(), resume_df, self.csv_file.get(), self.output_file.get())
            messagebox.showinfo("Success", "Excel file generated successfully!")
        except Exception as e:
            messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    root = Tk()
    app = App(root)
    root.mainloop()
